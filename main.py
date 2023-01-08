from openpyxl import Workbook
import openpyxl
wb = openpyxl.load_workbook("import_sheet_1.xlsx")

# wb.sheetnames

#getting a particular sheet
sheet1 = wb["Sponsored Product Keyword Repor"]
# print(sheet1.cell(row=1, column=3).value)
f = open('fatmug.csv','w')
f.write('Date,Portfolio name, Campaign Name,Targeting,spend,click,sale,average acos,average cpc,final bid \n')
col = sheet1.max_column
row = sheet1.max_row

t_spend = 0
t_click = 0
t_sale = 0

# for k in range(1, row+1):
for i in range(1, row+1):
    # Match Type
    v = sheet1.cell( row=i, column=7).value
    # Exact
    if v == 'EXACT':
        # --------- Date 
        date = sheet1.cell( row=i, column=1).value
        # print(date)
        # --------- Portfolio name 
        portfolio = sheet1.cell( row=i, column=2).value
        # print(portfolio)
        # ---------Campaign Name
        campaign = sheet1.cell( row=i, column=4).value
        # print(campaign)
        # ---------Targeting 
        targeting = sheet1.cell( row=i, column=6).value
        # print(targeting)
        # ---------spend
        spend = sheet1.cell( row=i, column=12).value
        # print(spend)
        t_spend = t_spend+spend
        # ------- Click
        click = sheet1.cell( row=i, column=9).value
        # print(click)
        t_click = t_click+click
        # --------- sale
        sale = sheet1.cell( row=i, column=15).value
        # print(sale)
        t_sale = t_sale+sale
        # print(f'{portfolio},{campaign},{targeting},{spend},{click}, {sale}\n')
        f.write(f'{date},{portfolio},{campaign},{targeting},{spend},{click}, {sale}\n')
#  average_acos(in %) = total spend / total sales
a = t_spend
b = t_sale
average_acos = ( a / b ) if b != 0 else 0
# print('average_acos', average_acos)

# average_cpc = total_spend/total_clicks
a = t_spend
b = t_click
average_cpc = ( a / b ) if b != 0 else 0
# print('average_cpc', average_cpc)

if average_acos < 10:
    final_bid = 10 / average_acos * average_cpc
    # print(final_bid)
print(f'{average_acos},{average_cpc}, {final_bid}\n')
f.write(f' , , , , , , , {average_acos}, {average_cpc}, {final_bid} \n')

print('\n')
f.close()














# expwb = openpyxl.load_workbook("export_sheet_1.xlsx")

# sheetex = exwb.sheetnames
# print(sheetex)

#getting a particular sheet
# sheetexp = expwb["Sponsored Products Campaigns"]
# print(sheetexp.cell(row=1, column=3).value)
# sheetexp=sheetexp.active


# sheetexp['A3'] = 'qwertyui'


# sheetexp.cell(row=12,column=2).value=5
# wb.save('export_sheet_1.xlsx')

